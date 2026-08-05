"""
检查模板和生成的Excel文件
"""
import os
from openpyxl import load_workbook

work_dir = r"c:\Users\m1316\Desktop\值班记录表"

print("=" * 60)
print("检查模板文件")
print("=" * 60)

template_path = os.path.join(work_dir, "docs", "电子值班表模板.xlsx")
print(f"模板文件路径: {template_path}")
print(f"模板存在: {os.path.exists(template_path)}")
print()

if os.path.exists(template_path):
    try:
        wb = load_workbook(template_path)
        print(f"模板工作表: {wb.sheetnames}")
        print()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"工作表: {sheet_name}")
            print(f"  A1: {repr(ws['A1'].value)}")
            print(f"  A2: {repr(ws['A2'].value)}")
            print(f"  A3: {repr(ws['A3'].value)}")
            print(f"  A4: {repr(ws['A4'].value)}")
            print(f"  B4: {repr(ws['B4'].value)}")

            # 检查验证
            try:
                validations = ws.data_validations.dataValidation
                print(f"  数据验证数量: {len(validations)}")
                for dv in validations:
                    print(f"    - 范围: {dv.sqref}, 公式: {dv.formula1}")
            except Exception as e:
                print(f"  检查验证失败: {e}")
            print()
        wb.close()
    except Exception as e:
        print(f"读取模板失败: {e}")

print("=" * 60)
print("检查生成的Excel文件")
print("=" * 60)

excel_files = [f for f in os.listdir(work_dir) if f.startswith("值班记录_") and f.endswith(".xlsx")]
if excel_files:
    target_file = os.path.join(work_dir, excel_files[0])
    print(f"文件: {target_file}")
    print()

    try:
        wb = load_workbook(target_file)
        print(f"工作表: {wb.sheetnames}")
        print()

        for sheet_name in wb.sheetnames:
            if len(sheet_name) == 10 and sheet_name[4] == '-' and sheet_name[7] == '-':
                ws = wb[sheet_name]
                print(f"日期工作表: {sheet_name}")
                print(f"  A1: {repr(ws['A1'].value)}")
                print(f"  A2: {repr(ws['A2'].value)}")
                print(f"  A3: {repr(ws['A3'].value)}")
                print(f"  A4: {repr(ws['A4'].value)}")

                try:
                    validations = ws.data_validations.dataValidation
                    print(f"  数据验证数量: {len(validations)}")
                    for dv in validations:
                        print(f"    - 范围: {dv.sqref}, 公式: {dv.formula1}")
                except Exception as e:
                    print(f"  检查验证失败: {e}")
        wb.close()
    except Exception as e:
        print(f"读取文件失败: {e}")
else:
    print("未找到生成的Excel文件")

print("=" * 60)