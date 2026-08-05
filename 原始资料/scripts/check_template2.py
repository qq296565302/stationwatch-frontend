"""
检查还原后的模板
"""
import os
from openpyxl import load_workbook

work_dir = r"c:\Users\m1316\Desktop\值班记录表"
template_path = os.path.join(work_dir, "docs", "电子值班表模板.xlsx")

print(f"模板路径: {template_path}")
print(f"存在: {os.path.exists(template_path)}")
print()

if os.path.exists(template_path):
    wb = load_workbook(template_path)
    print(f"工作表: {wb.sheetnames}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"=== {sheet_name} ===")
        print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")

        # 打印前22行
        for row in range(1, min(23, ws.max_row + 1)):
            for col in range(1, min(12, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value is not None:
                    print(f"  {cell.coordinate}: {repr(cell.value)}")
        print()
    wb.close()