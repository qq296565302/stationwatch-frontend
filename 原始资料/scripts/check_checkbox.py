"""
检查复选框位置和合并单元格
"""
import os
from openpyxl import load_workbook
import zipfile

work_dir = r"c:\Users\m1316\Desktop\值班记录表"
excel_files = [f for f in os.listdir(work_dir) if f.startswith("值班记录_") and f.endswith(".xlsx")]

target_file = os.path.join(work_dir, excel_files[0])
print(f"文件: {target_file}")
print()

# 使用openpyxl检查
wb = load_workbook(target_file)
for sheet_name in wb.sheetnames:
    if len(sheet_name) == 10 and sheet_name[4] == '-' and sheet_name[7] == '-':
        ws = wb[sheet_name]
        print(f"工作表: {sheet_name}")
        print(f"合并单元格: {list(ws.merged_cells.ranges)}")
        print()
        break
wb.close()

# 使用zipfile直接查看xml中的复选框位置
print("=" * 60)
print("查看XML中的控件位置")
print("=" * 60)

with zipfile.ZipFile(target_file, 'r') as z:
    # 查看所有xml文件
    names = z.namelist()
    sheet_files = [n for n in names if 'sheet' in n.lower() and n.endswith('.xml')]
    print(f"Sheet文件: {sheet_files}")
    print()

    # 查看第一个sheet的xml
    for sheet_file in sheet_files:
        if 'sheet1.xml' in sheet_file or 'sheet2.xml' in sheet_file:
            with z.open(sheet_file) as f:
                content = f.read().decode('utf-8', errors='ignore')
                # 查找控件
                if 'checkbox' in content.lower() or 'formControl' in content or 'checkBox' in content:
                    print(f"=== {sheet_file} 包含复选框 ===")

                    # 查找所有控件位置
                    import re
                    # 查找 from/to 标记
                    from_matches = re.findall(r'<x:from>(.*?)</x:from>', content, re.DOTALL)
                    for m in from_matches:
                        print(f"控件位置: {m}")
                    print()
                    break