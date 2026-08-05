"""
检查生成的Excel文件
"""
import os
from openpyxl import load_workbook

work_dir = r"c:\Users\m1316\Desktop\值班记录表"
excel_files = [f for f in os.listdir(work_dir) if f.startswith("值班记录_") and f.endswith(".xlsx")]

if not excel_files:
    print("未找到Excel文件")
    exit(1)

target_file = os.path.join(work_dir, excel_files[0])
wb = load_workbook(target_file)

for sheet_name in wb.sheetnames:
    if len(sheet_name) == 10 and sheet_name[4] == '-' and sheet_name[7] == '-':
        ws = wb[sheet_name]
        print(f"工作表: {sheet_name}")
        for row in range(1, 6):
            for col in range(1, 12):
                cell = ws.cell(row, col)
                if cell.value is not None:
                    print(f"  {cell.coordinate}: {repr(cell.value)}")
            print()
        break

wb.close()